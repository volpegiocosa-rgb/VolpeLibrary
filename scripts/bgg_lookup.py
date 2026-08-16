#!/usr/bin/env python3
"""Arricchisce il registro giochi (xlsx) con i dati scaricati da BoardGameGeek (BGG).

Legge `legacy/Registro_giochi.xlsx` (foglio "Tabella", colonne Gioco / Posizione)
e per ogni gioco cerca su BGG giocatori minimi, massimi, durata, peso
(complessita'), tipo (subdomain BGG: Abstract Game, Strategy Game, Family
Game, ...), categorie, meccaniche e famiglie. Scrive un JSON completo (dati
del registro + dati BGG) in `output/catalogo_bgg.json`. La cartella `legacy/`
viene solo letta, mai scritta.

Uso:
    export BGG_API_TOKEN="il-tuo-token"
    python3 scripts/bgg_lookup.py                       # elabora tutto il registro
    python3 scripts/bgg_lookup.py --limit 5              # solo i primi 5, per provare
    python3 scripts/bgg_lookup.py --resume               # riprende saltando i giochi gia' in output
    python3 scripts/bgg_lookup.py --game "Catan" --json  # ricerca puntuale di un singolo gioco

NOTE:
- BGG richiede ora un token di autorizzazione (Bearer) per la XML API2, ottenuto
  tramite registrazione approvata da BGG. Finche' il token non e' attivo, ogni
  chiamata restituira' 401/403 (vedi BggAuthError piu' sotto).
- Questo modulo non e' ancora stato testato contro l'API reale: non appena il
  token e' disponibile va verificato prima con --game su un gioco noto (es.
  "Catan") e poi con --limit su poche righe del registro.
"""

from __future__ import annotations

import argparse
import json
import os
import random
import sys
import time
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import requests

BASE_URL = "https://boardgamegeek.com/xmlapi2"
SEARCH_URL = f"{BASE_URL}/search"
THING_URL = f"{BASE_URL}/thing"

LINK_TYPE_CATEGORY = "boardgamecategory"
LINK_TYPE_MECHANIC = "boardgamemechanic"
LINK_TYPE_FAMILY = "boardgamefamily"
RANK_TYPE_SUBDOMAIN = "family"  # es. Abstract/Strategy/Family/Party/Thematic/War Game Rank

REQUEST_TIMEOUT = 15  # secondi
QUEUE_RETRY_DELAY = 5  # secondi, per il retry su HTTP 202 (richiesta in coda)
RATE_LIMIT_INITIAL_DELAY = 5  # secondi, per il retry su HTTP 429 (rate limit); raddoppia ad ogni tentativo
RATE_LIMIT_MAX_DELAY = 120  # secondi
MAX_RETRIES = 10

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX_PATH = REPO_ROOT / "legacy" / "Registro_giochi.xlsx"
DEFAULT_SHEET_NAME = "Tabella"
DEFAULT_OUTPUT_PATH = REPO_ROOT / "output" / "catalogo_bgg.json"
DEFAULT_REQUEST_DELAY = 5.0  # secondi tra una scheda e l'altra, per rispettare i rate limit BGG


def _jitter(seconds: float, spread: float) -> float:
    """Aggiunge un jitter casuale a un delay, per evitare un pattern di richieste troppo regolare."""
    return seconds + random.uniform(0, spread)


class BggError(Exception):
    """Errore generico nella comunicazione con BGG."""


class BggAuthError(BggError):
    """Token mancante, non valido o non ancora approvato da BGG."""


class BggNotFoundError(BggError):
    """Nessun gioco trovato per il nome dato."""


@dataclass
class BggSearchResult:
    bgg_id: str
    nome: str
    anno: str | None


@dataclass
class BggGameInfo:
    bgg_id: str
    nome: str
    giocatori_min: int | None
    giocatori_max: int | None
    durata_minuti: int | None
    peso: float | None  # "weight" BGG: complessita' media, scala 1-5
    tipo: list[str]  # subdomain BGG: Abstract Game, Strategy Game, Family Game, ...
    categorie: list[str]
    meccaniche: list[str]
    famiglie: list[str]


@dataclass
class RegistroEntry:
    gioco: str
    posizione: str


def _auth_headers(token: str) -> dict:
    return {
        "Authorization": f"Bearer {token}",
        "User-Agent": "VolpeGiocosa-DataFetcher/1.0 (+https://volpegiocosa.it)",
    }


def _request_with_retry(url: str, params: dict, token: str) -> ET.Element:
    """Esegue una GET verso la XML API2 gestendo auth, coda asincrona e rate limit.

    BGG puo' rispondere 202 mentre prepara i dati (in particolare su /thing): in
    quel caso va ripetuta la richiesta dopo una breve pausa fissa. Puo' anche
    rispondere 429 (troppe richieste): in quel caso si aspetta l'header
    Retry-After se presente, altrimenti un backoff esponenziale, stampando
    sempre a schermo cosa sta succedendo cosi' l'utente vede perche' il run
    rallenta invece di sembrare bloccato.
    """
    headers = _auth_headers(token)
    rate_limit_delay = RATE_LIMIT_INITIAL_DELAY

    for attempt in range(MAX_RETRIES):
        try:
            response = requests.get(
                url, params=params, headers=headers, timeout=REQUEST_TIMEOUT
            )
        except requests.RequestException as exc:
            attesa = _jitter(RATE_LIMIT_INITIAL_DELAY, spread=1.0)
            print(f"  errore di rete, ritento tra {attesa:.1f}s: {exc}", file=sys.stderr)
            time.sleep(attesa)
            continue

        if response.status_code in (401, 403):
            raise BggAuthError(
                "BGG ha rifiutato il token (status "
                f"{response.status_code}). Verifica che BGG_API_TOKEN sia "
                "corretto e che la registrazione sia stata approvata."
            )

        if response.status_code == 202:
            # Richiesta accettata ma dati non ancora pronti: ritenta.
            attesa = _jitter(QUEUE_RETRY_DELAY, spread=1.0)
            print(
                f"  BGG sta ancora preparando i dati (202), riprovo tra {attesa:.1f}s...",
                file=sys.stderr,
            )
            time.sleep(attesa)
            continue

        if response.status_code == 429:
            retry_after = response.headers.get("Retry-After")
            if retry_after and retry_after.replace(".", "", 1).isdigit():
                attesa = _jitter(float(retry_after), spread=1.0)
            else:
                attesa = _jitter(rate_limit_delay, spread=rate_limit_delay * 0.5)
                rate_limit_delay = min(rate_limit_delay * 2, RATE_LIMIT_MAX_DELAY)
            print(
                f"  RATE LIMIT (429) da BGG: aspetto {attesa:.1f}s e ritento "
                f"(tentativo {attempt + 1}/{MAX_RETRIES})...",
                file=sys.stderr,
            )
            time.sleep(attesa)
            continue

        if not response.ok:
            raise BggError(
                f"BGG ha risposto con status {response.status_code} per {url}"
            )

        try:
            return ET.fromstring(response.content)
        except ET.ParseError as exc:
            raise BggError(f"Risposta XML non valida da BGG: {exc}") from exc

    raise BggError(
        f"Troppi tentativi falliti (coda/rate-limit/rete) per {url}, riprova piu' tardi."
    )


def search_game(nome: str, token: str) -> list[BggSearchResult]:
    """Cerca un gioco per nome e restituisce i candidati trovati su BGG."""
    root = _request_with_retry(
        SEARCH_URL, {"query": nome, "type": "boardgame"}, token
    )

    risultati: list[BggSearchResult] = []
    for item in root.findall("item"):
        bgg_id = item.get("id")
        if bgg_id is None:
            continue
        name_el = item.find("name")
        year_el = item.find("yearpublished")
        risultati.append(
            BggSearchResult(
                bgg_id=bgg_id,
                nome=name_el.get("value") if name_el is not None else nome,
                anno=year_el.get("value") if year_el is not None else None,
            )
        )
    return risultati


def _pick_best_match(nome: str, candidati: list[BggSearchResult]) -> BggSearchResult:
    nome_normalizzato = nome.strip().casefold()
    for candidato in candidati:
        if candidato.nome.strip().casefold() == nome_normalizzato:
            return candidato
    # Nessuna corrispondenza esatta: BGG ordina i risultati per rilevanza,
    # quindi il primo candidato resta la scelta migliore disponibile.
    return candidati[0]


def _link_values(item: ET.Element, link_type: str) -> list[str]:
    return [
        link.get("value")
        for link in item.findall("link")
        if link.get("type") == link_type and link.get("value")
    ]


def _subdomain_types(item: ET.Element) -> list[str]:
    """Estrae il "tipo" BGG (Abstract/Strategy/Family/Party/Thematic/War/...) dai rank."""
    ranks = item.find("statistics/ratings/ranks")
    if ranks is None:
        return []
    tipi = []
    for rank in ranks.findall("rank"):
        if rank.get("type") != RANK_TYPE_SUBDOMAIN:
            continue
        friendly = rank.get("friendlyname") or rank.get("name")
        if friendly:
            tipi.append(friendly.removesuffix(" Rank").strip())
    return tipi


def get_game_details(bgg_id: str, token: str) -> BggGameInfo:
    """Scarica i dettagli (giocatori, durata, peso, tipo, categorie, meccaniche, famiglie) per un id BGG."""
    root = _request_with_retry(THING_URL, {"id": bgg_id, "stats": 1}, token)

    item = root.find("item")
    if item is None:
        raise BggNotFoundError(f"Nessun dettaglio trovato per l'id BGG {bgg_id}")

    def _int_value(tag: str) -> int | None:
        el = item.find(tag)
        if el is None or el.get("value") in (None, "0", ""):
            return None
        return int(el.get("value"))

    name_el = item.find("name")
    nome = name_el.get("value") if name_el is not None else bgg_id

    peso_el = item.find("statistics/ratings/averageweight")
    peso = None
    if peso_el is not None and peso_el.get("value") not in (None, "0", ""):
        peso = round(float(peso_el.get("value")), 2)

    return BggGameInfo(
        bgg_id=bgg_id,
        nome=nome,
        giocatori_min=_int_value("minplayers"),
        giocatori_max=_int_value("maxplayers"),
        durata_minuti=_int_value("playingtime"),
        peso=peso,
        tipo=_subdomain_types(item),
        categorie=_link_values(item, LINK_TYPE_CATEGORY),
        meccaniche=_link_values(item, LINK_TYPE_MECHANIC),
        famiglie=_link_values(item, LINK_TYPE_FAMILY),
    )


def lookup_game(nome: str, token: str) -> BggGameInfo:
    """Cerca un gioco per nome su BGG e ne scarica i dettagli."""
    candidati = search_game(nome, token)
    if not candidati:
        raise BggNotFoundError(f'Nessun gioco trovato su BGG per "{nome}"')

    scelto = _pick_best_match(nome, candidati)
    return get_game_details(scelto.bgg_id, token)


def read_registro(
    xlsx_path: Path, sheet_name: str = DEFAULT_SHEET_NAME
) -> list[RegistroEntry]:
    """Legge il registro giochi dall'xlsx compilato a mano (colonne Gioco/Posizione)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise BggError(
            f"Foglio '{sheet_name}' non trovato in {xlsx_path}. "
            f"Fogli disponibili: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    righe = ws.iter_rows(values_only=True)
    intestazione = [str(h).strip() if h else "" for h in next(righe)]

    colonne_richieste = ("Gioco", "Posizione")
    try:
        indici = {col: intestazione.index(col) for col in colonne_richieste}
    except ValueError as exc:
        raise BggError(
            f"Intestazioni attese {colonne_richieste} non trovate nel foglio "
            f"'{sheet_name}'. Trovate: {intestazione}"
        ) from exc

    voci: list[RegistroEntry] = []
    for riga in righe:
        gioco = riga[indici["Gioco"]]
        if not gioco:
            continue
        posizione = riga[indici["Posizione"]] or ""
        voci.append(
            RegistroEntry(
                gioco=str(gioco).strip(),
                posizione=str(posizione).strip(),
            )
        )
    return voci


def build_catalog(
    xlsx_path: Path,
    token: str,
    output_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    delay: float = DEFAULT_REQUEST_DELAY,
    limit: int | None = None,
    resume: bool = False,
) -> list[dict]:
    """Legge il registro xlsx, interroga BGG per ogni gioco e scrive il JSON completo.

    Salva il JSON dopo ogni gioco elaborato, cosi' un'interruzione (rete, rate
    limit, Ctrl+C) non fa perdere il lavoro gia' fatto. Con --resume i giochi
    gia' presenti in output_path e recuperati CON SUCCESSO vengono saltati;
    quelli che in precedenza avevano dato errore (es. rate limit transitorio)
    vengono invece ritentati.
    """
    voci = read_registro(xlsx_path, sheet_name)
    if limit is not None:
        voci = voci[:limit]

    # Chiave (Gioco, Posizione) invece del solo nome: il registro puo' avere lo
    # stesso titolo su piu' righe (copie fisiche diverse in posizioni diverse),
    # e deduplicare per solo nome ne perderebbe silenziosamente le altre copie.
    risultati: list[dict] = []
    gia_elaborati: set[tuple[str, str]] = set()
    if resume and output_path.exists():
        with open(output_path, encoding="utf-8") as f:
            precedenti = json.load(f)
        da_ritentare = 0
        for voce in precedenti:
            if voce.get("bgg_errore"):
                da_ritentare += 1
                continue
            risultati.append(voce)
            gia_elaborati.add((voce["Gioco"], voce["Posizione"]))
        print(
            f"Ripresa: {len(gia_elaborati)} gia' completati, {da_ritentare} da ritentare.",
            file=sys.stderr,
        )

    for i, voce in enumerate(voci, start=1):
        if (voce.gioco, voce.posizione) in gia_elaborati:
            continue

        print(f"[{i}/{len(voci)}] {voce.gioco} ...", file=sys.stderr)
        record = {
            "Gioco": voce.gioco,
            "Posizione": voce.posizione,
            "bgg_id": None,
            "giocatori_min": None,
            "giocatori_max": None,
            "durata_minuti": None,
            "peso": None,
            "tipo": [],
            "categorie": [],
            "meccaniche": [],
            "famiglie": [],
            "bgg_errore": None,
        }

        try:
            info = lookup_game(voce.gioco, token)
        except BggAuthError:
            # Token non valido/non approvato: non ha senso continuare il batch.
            raise
        except BggError as exc:
            record["bgg_errore"] = str(exc)
            print(f"  attenzione: {exc}", file=sys.stderr)
        else:
            record.update(
                bgg_id=info.bgg_id,
                giocatori_min=info.giocatori_min,
                giocatori_max=info.giocatori_max,
                durata_minuti=info.durata_minuti,
                peso=info.peso,
                tipo=info.tipo,
                categorie=info.categorie,
                meccaniche=info.meccaniche,
                famiglie=info.famiglie,
            )

        risultati.append(record)

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(risultati, f, ensure_ascii=False, indent=2)

        if i < len(voci):
            time.sleep(_jitter(delay, spread=delay * 0.5))

    ordine = {(voce.gioco, voce.posizione): i for i, voce in enumerate(voci)}
    risultati.sort(key=lambda r: ordine.get((r["Gioco"], r["Posizione"]), len(voci)))

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(risultati, f, ensure_ascii=False, indent=2)

    falliti = [r for r in risultati if r.get("bgg_errore")]
    print(
        f"Riepilogo: {len(risultati)} totali, "
        f"{len(risultati) - len(falliti)} riusciti, {len(falliti)} falliti.",
        file=sys.stderr,
    )
    if falliti:
        print("Giochi falliti:", file=sys.stderr)
        for r in falliti:
            print(f"  - {r['Gioco']}: {r['bgg_errore']}", file=sys.stderr)

    return risultati


def _print_game_info(info: BggGameInfo, as_json: bool) -> None:
    if as_json:
        print(
            json.dumps(
                {
                    "bgg_id": info.bgg_id,
                    "nome": info.nome,
                    "giocatori_min": info.giocatori_min,
                    "giocatori_max": info.giocatori_max,
                    "durata_minuti": info.durata_minuti,
                    "peso": info.peso,
                    "tipo": info.tipo,
                    "categorie": info.categorie,
                    "meccaniche": info.meccaniche,
                    "famiglie": info.famiglie,
                },
                ensure_ascii=False,
                indent=2,
            )
        )
    else:
        print(f"Gioco: {info.nome} (BGG id {info.bgg_id})")
        print(f"Giocatori: {info.giocatori_min}-{info.giocatori_max}")
        print(f"Durata: {info.durata_minuti} minuti")
        print(f"Peso: {info.peso}")
        print(f"Tipo: {', '.join(info.tipo)}")
        print(f"Categorie: {', '.join(info.categorie)}")
        print(f"Meccaniche: {', '.join(info.meccaniche)}")
        print(f"Famiglie: {', '.join(info.famiglie)}")


def _parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX_PATH,
        help=f"Percorso del registro xlsx (default: {DEFAULT_XLSX_PATH})",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
        help=f"Nome del foglio da leggere (default: {DEFAULT_SHEET_NAME})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"Percorso del JSON di output (default: {DEFAULT_OUTPUT_PATH})",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=DEFAULT_REQUEST_DELAY,
        help="Secondi di pausa tra un gioco e l'altro (default: %(default)s)",
    )
    parser.add_argument(
        "--limit", type=int, default=None, help="Elabora solo le prime N righe (utile per provare)"
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Riprende un'elaborazione precedente saltando i giochi gia' in --output",
    )
    parser.add_argument(
        "--game",
        default=None,
        help="Ignora l'xlsx e cerca un singolo gioco per nome",
    )
    parser.add_argument(
        "--json", action="store_true", help="Con --game, stampa il risultato come JSON"
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv if argv is not None else sys.argv[1:])

    token = os.environ.get("BGG_API_TOKEN")
    if not token:
        print(
            "Errore: variabile d'ambiente BGG_API_TOKEN non impostata.",
            file=sys.stderr,
        )
        return 1

    if args.game:
        try:
            info = lookup_game(args.game, token)
        except BggError as exc:
            print(f"Errore: {exc}", file=sys.stderr)
            return 1
        _print_game_info(info, as_json=args.json)
        return 0

    try:
        risultati = build_catalog(
            xlsx_path=args.xlsx,
            token=token,
            output_path=args.output,
            sheet_name=args.sheet,
            delay=args.delay,
            limit=args.limit,
            resume=args.resume,
        )
    except BggError as exc:
        print(f"Errore: {exc}", file=sys.stderr)
        return 1

    print(f"Scritte {len(risultati)} voci in {args.output}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
